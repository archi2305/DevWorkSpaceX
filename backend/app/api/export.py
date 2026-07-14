import io
import csv
import json
import zipfile
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, status, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.dependencies.db import get_db
from app.dependencies.auth import get_current_user
from app.models.user import User
from app.models.project import Project
from app.models.task import Task
from app.models.document import Document
from app.models.file_asset import FileAsset
from app.models.activity import ActivityLog
from app.models.label import Label
from app.models.milestone import Milestone
from app.models.sprint import Sprint

router = APIRouter(prefix="/workspace/export", tags=["Workspace Export"])

def _serialize_datetime(dt: Optional[datetime]) -> Optional[str]:
    """Serialize datetime to ISO format string."""
    return dt.isoformat() if dt else None

def _get_project_data(project: Project) -> dict:
    """Extract comprehensive project data."""
    return {
        "id": str(project.id),
        "name": project.name,
        "slug": project.slug,
        "description": project.description,
        "icon": project.icon,
        "cover_image": project.cover_image,
        "color": project.color,
        "status": project.status,
        "priority": project.priority,
        "progress": project.progress,
        "is_favorite": project.is_favorite,
        "is_pinned": project.is_pinned,
        "is_archived": project.is_archived,
        "is_template": project.is_template,
        "is_deleted": project.is_deleted,
        "deleted_at": _serialize_datetime(project.deleted_at),
        "owner_id": str(project.owner_id) if project.owner_id else None,
        "workspace_id": str(project.workspace_id) if project.workspace_id else None,
        "visibility": project.visibility,
        "due_date": _serialize_datetime(project.due_date),
        "kanban_columns": project.kanban_columns,
        "created_at": _serialize_datetime(project.created_at),
        "updated_at": _serialize_datetime(project.updated_at)
    }

def _get_task_data(task: Task) -> dict:
    """Extract comprehensive task data."""
    return {
        "id": str(task.id),
        "title": task.title,
        "description": task.description,
        "status": task.status,
        "due_date": task.due_date,
        "priority": task.priority,
        "completed": task.completed,
        "assignee_id": str(task.assignee_id) if task.assignee_id else None,
        "project_id": str(task.project_id) if task.project_id else None,
        "sprint_id": str(task.sprint_id) if task.sprint_id else None,
        "story_points": task.story_points,
        "estimated_time": task.estimated_time,
        "is_archived": task.is_archived,
        "is_deleted": task.is_deleted,
        "deleted_at": _serialize_datetime(task.deleted_at),
        "attachments": task.attachments,
        "parent_id": str(task.parent_id) if task.parent_id else None,
        "milestone_id": str(task.milestone_id) if task.milestone_id else None,
        "release_id": str(task.release_id) if task.release_id else None,
        "created_at": _serialize_datetime(task.created_at),
        "updated_at": _serialize_datetime(task.updated_at)
    }

def _get_document_data(document: Document) -> dict:
    """Extract comprehensive document data."""
    return {
        "id": str(document.id),
        "title": document.title,
        "content": document.content,
        "project_id": str(document.project_id) if document.project_id else None,
        "parent_id": str(document.parent_id) if document.parent_id else None,
        "is_favorite": document.is_favorite,
        "version": document.version,
        "owner_id": str(document.owner_id) if document.owner_id else None,
        "created_at": _serialize_datetime(document.created_at),
        "updated_at": _serialize_datetime(document.updated_at)
    }

def _get_file_data(file_asset: FileAsset) -> dict:
    """Extract comprehensive file data."""
    return {
        "id": str(file_asset.id),
        "name": file_asset.name,
        "file_path": file_asset.file_path,
        "mime_type": file_asset.mime_type,
        "size": file_asset.size,
        "is_folder": file_asset.is_folder,
        "project_id": str(file_asset.project_id) if file_asset.project_id else None,
        "parent_id": str(file_asset.parent_id) if file_asset.parent_id else None,
        "owner_id": str(file_asset.owner_id) if file_asset.owner_id else None,
        "created_at": _serialize_datetime(file_asset.created_at),
        "updated_at": _serialize_datetime(file_asset.updated_at)
    }

def _get_activity_data(activity: ActivityLog) -> dict:
    """Extract comprehensive activity data."""
    return {
        "id": str(activity.id),
        "user_id": str(activity.user_id) if activity.user_id else None,
        "action": activity.action,
        "details": activity.details,
        "target_type": activity.target_type,
        "target_name": activity.target_name,
        "created_at": _serialize_datetime(activity.created_at)
    }

def _get_workspace_data(db: Session, data_types: Optional[List[str]] = None):
    """Get workspace data based on selected data types."""
    if data_types is None:
        data_types = ["projects", "tasks", "documents", "files", "activities"]
    
    result = {}
    
    if "projects" in data_types:
        projects = db.query(Project).all()
        result["projects"] = [_get_project_data(p) for p in projects]
    
    if "tasks" in data_types:
        tasks = db.query(Task).all()
        result["tasks"] = [_get_task_data(t) for t in tasks]
    
    if "documents" in data_types:
        documents = db.query(Document).all()
        result["documents"] = [_get_document_data(d) for d in documents]
    
    if "files" in data_types:
        files = db.query(FileAsset).all()
        result["files"] = [_get_file_data(f) for f in files]
    
    if "activities" in data_types:
        activities = db.query(ActivityLog).all()
        result["activities"] = [_get_activity_data(a) for a in activities]
    
    if "labels" in data_types:
        labels = db.query(Label).all()
        result["labels"] = [
            {
                "id": str(l.id),
                "name": l.name,
                "color": l.color,
                "workspace_id": str(l.workspace_id) if l.workspace_id else None,
                "created_at": _serialize_datetime(l.created_at)
            }
            for l in labels
        ]
    
    if "milestones" in data_types:
        milestones = db.query(Milestone).all()
        result["milestones"] = [
            {
                "id": str(m.id),
                "name": m.name,
                "description": m.description,
                "project_id": str(m.project_id) if m.project_id else None,
                "status": m.status,
                "due_date": _serialize_datetime(m.due_date),
                "created_at": _serialize_datetime(m.created_at)
            }
            for m in milestones
        ]
    
    if "sprints" in data_types:
        sprints = db.query(Sprint).all()
        result["sprints"] = [
            {
                "id": str(s.id),
                "name": s.name,
                "description": s.description,
                "project_id": str(s.project_id) if s.project_id else None,
                "status": s.status,
                "start_date": _serialize_datetime(s.start_date),
                "end_date": _serialize_datetime(s.end_date),
                "created_at": _serialize_datetime(s.created_at)
            }
            for s in sprints
        ]
    
    return result

def _generate_csv(data: list, fieldnames: Optional[list] = None) -> str:
    """Generate CSV from data list."""
    if not data:
        return ""
    
    output = io.StringIO()
    if fieldnames is None:
        fieldnames = list(data[0].keys())
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in data:
        filtered_row = {k: str(v) if v is not None else "" for k, v in row.items() if k in fieldnames}
        writer.writerow(filtered_row)
    return output.getvalue()

def _generate_multi_csv(data: dict) -> str:
    """Generate combined CSV with section headers for multiple data types."""
    output = io.StringIO()
    
    for section_name, rows in data.items():
        if not rows:
            continue
        
        output.write(f"# {section_name.upper()}\n")
        output.write(_generate_csv(rows))
        output.write("\n")
    
    return output.getvalue()

def _generate_excel(data: dict) -> bytes:
    """Generate Excel file with multiple sheets using openpyxl."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for sheet_name, rows in data.items():
        if not rows:
            continue
        
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars
        
        # Write headers
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row.get(key)
                if value is None:
                    value = ""
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value)
                else:
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def _generate_markdown(data: dict) -> str:
    """Generate comprehensive Markdown export."""
    md = "# Workspace Export\n\n"
    md += f"*Generated on {datetime.utcnow().isoformat()}*\n\n"
    md += "---\n\n"
    
    # Projects Section
    if "projects" in data and data["projects"]:
        md += "## Projects\n\n"
        md += "| ID | Name | Status | Priority | Progress | Created | Updated |\n"
        md += "| --- | --- | --- | --- | --- | --- | --- |\n"
        for p in data["projects"]:
            md += f"| {p['id'][:8]}... | {p['name']} | {p['status']} | {p['priority']} | {p['progress']}% | {p['created_at'][:10]} | {p['updated_at'][:10]} |\n"
        md += "\n"
    
    # Tasks Section
    if "tasks" in data and data["tasks"]:
        md += "## Tasks\n\n"
        md += "| ID | Title | Status | Priority | Assignee | Due Date | Created |\n"
        md += "| --- | --- | --- | --- | --- | --- | --- |\n"
        for t in data["tasks"]:
            md += f"| {t['id'][:8]}... | {t['title'][:50]} | {t['status']} | {t['priority']} | {t['assignee_id'][:8] if t['assignee_id'] else 'N/A'}... | {t['due_date'] or 'N/A'} | {t['created_at'][:10]} |\n"
        md += "\n"
    
    # Documents Section
    if "documents" in data and data["documents"]:
        md += "## Documents\n\n"
        for d in data["documents"]:
            md += f"### {d['title']}\n\n"
            md += f"*ID: {d['id']} | Version: {d['version']} | Created: {d['created_at']}*\n\n"
            content = d['content'] or 'No content'
            if len(content) > 500:
                content = content[:500] + "..."
            md += f"{content}\n\n"
        md += "\n"
    
    # Files Section
    if "files" in data and data["files"]:
        md += "## Files\n\n"
        md += "| ID | Name | Type | Size | Folder | Created |\n"
        md += "| --- | --- | --- | --- | --- | --- |\n"
        for f in data["files"]:
            size_mb = f"{f['size'] / (1024*1024):.2f} MB" if f['size'] else 'N/A'
            md += f"| {f['id'][:8]}... | {f['name'][:30]} | {f['mime_type'] or 'N/A'} | {size_mb} | {f['is_folder']} | {f['created_at'][:10]} |\n"
        md += "\n"
    
    # Activity Section
    if "activities" in data and data["activities"]:
        md += "## Activity Log\n\n"
        md += "| Date | Action | Details | Target |\n"
        md += "| --- | --- | --- | --- |\n"
        for a in data["activities"]:
            md += f"| {a['created_at'][:10]} | {a['action']} | {a['details'][:50]} | {a['target_type'] or 'N/A'} |\n"
        md += "\n"
    
    # Labels Section
    if "labels" in data and data["labels"]:
        md += "## Labels\n\n"
        for l in data["labels"]:
            md += f"- **{l['name']}** (Color: {l['color']})\n"
        md += "\n"
    
    # Milestones Section
    if "milestones" in data and data["milestones"]:
        md += "## Milestones\n\n"
        md += "| Name | Status | Due Date |\n"
        md += "| --- | --- | --- |\n"
        for m in data["milestones"]:
            md += f"| {m['name']} | {m['status']} | {m['due_date'] or 'N/A'} |\n"
        md += "\n"
    
    # Sprints Section
    if "sprints" in data and data["sprints"]:
        md += "## Sprints\n\n"
        md += "| Name | Status | Start | End |\n"
        md += "| --- | --- | --- | --- |\n"
        for s in data["sprints"]:
            md += f"| {s['name']} | {s['status']} | {s['start_date'] or 'N/A'} | {s['end_date'] or 'N/A'} |\n"
        md += "\n"
    
    md += "---\n\n*End of Export*\n"
    return md

@router.get("")
def export_workspace(
    format: str = Query("json", description="Export format: json, csv, excel, markdown, zip"),
    data_types: Optional[str] = Query(None, description="Comma-separated data types: projects,tasks,documents,files,activities,labels,milestones,sprints"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export workspace data in various formats."""
    # Parse data types
    type_list = None
    if data_types:
        type_list = [t.strip().lower() for t in data_types.split(",")]
        valid_types = ["projects", "tasks", "documents", "files", "activities", "labels", "milestones", "sprints"]
        type_list = [t for t in type_list if t in valid_types]
    
    data = _get_workspace_data(db, type_list)
    
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    
    if format == "json":
        json_data = json.dumps(data, indent=2, default=str)
        return StreamingResponse(
            io.BytesIO(json_data.encode()),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=workspace_export_{timestamp}.json"}
        )
        
    elif format == "csv":
        csv_str = _generate_multi_csv(data)
        return StreamingResponse(
            io.BytesIO(csv_str.encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=workspace_export_{timestamp}.csv"}
        )
        
    elif format == "excel":
        excel_data = _generate_excel(data)
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=workspace_export_{timestamp}.xlsx"}
        )
        
    elif format == "markdown":
        md_str = _generate_markdown(data)
        return StreamingResponse(
            io.BytesIO(md_str.encode()),
            media_type="text/markdown",
            headers={"Content-Disposition": f"attachment; filename=workspace_export_{timestamp}.md"}
        )
        
    elif format == "zip":
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Add JSON files for each data type
            for section_name, rows in data.items():
                if rows:
                    zip_file.writestr(f"{section_name}.json", json.dumps(rows, indent=2, default=str))
                    zip_file.writestr(f"{section_name}.csv", _generate_csv(rows))
            
            # Add comprehensive markdown summary
            zip_file.writestr("README.md", _generate_markdown(data))
            
            # Add full export as single JSON
            zip_file.writestr("full_export.json", json.dumps(data, indent=2, default=str))
            
        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename=workspace_export_{timestamp}.zip"}
        )
        
    else:
        raise HTTPException(status_code=400, detail="Unsupported export format. Supported formats: json, csv, excel, markdown, zip")
